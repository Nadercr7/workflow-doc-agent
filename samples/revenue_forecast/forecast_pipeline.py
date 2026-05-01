"""Monthly revenue forecast pipeline.

Reads the prior 24 months of actuals from `input_data.csv`, fits a simple
linear trend with seasonality, projects the next 6 months, and writes the
results into `monthly_revenue.xlsx` (3 sheets: Actuals, Forecast,
Sensitivity).

Designed to run as a scheduled monthly job. Can also be triggered ad hoc
when finance closes a new month and wants an updated outlook.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook

INPUT_CSV = Path(__file__).with_name("input_data.csv")
OUTPUT_XLSX = Path(__file__).with_name("monthly_revenue.xlsx")
FORECAST_HORIZON_MONTHS = 6


@dataclass(frozen=True, slots=True)
class MonthlyActual:
    period: str  # YYYY-MM
    revenue_usd: float
    new_customers: int
    churned_customers: int


def load_actuals(path: Path) -> list[MonthlyActual]:
    rows: list[MonthlyActual] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(
                MonthlyActual(
                    period=r["period"],
                    revenue_usd=float(r["revenue_usd"]),
                    new_customers=int(r["new_customers"]),
                    churned_customers=int(r["churned_customers"]),
                )
            )
    rows.sort(key=lambda x: x.period)
    return rows


def fit_linear_trend(values: list[float]) -> tuple[float, float]:
    """Return (slope, intercept) of a least-squares line through values."""
    n = len(values)
    if n < 2:
        return 0.0, values[0] if values else 0.0
    xs = list(range(n))
    mean_x = sum(xs) / n
    mean_y = sum(values) / n
    num = sum((xs[i] - mean_x) * (values[i] - mean_y) for i in range(n))
    den = sum((xs[i] - mean_x) ** 2 for i in range(n)) or 1.0
    slope = num / den
    intercept = mean_y - slope * mean_x
    return slope, intercept


def project_forecast(actuals: list[MonthlyActual], horizon: int) -> list[tuple[str, float]]:
    revenues = [a.revenue_usd for a in actuals]
    slope, intercept = fit_linear_trend(revenues)
    last_period = actuals[-1].period
    year, month = (int(p) for p in last_period.split("-"))

    out: list[tuple[str, float]] = []
    for i in range(1, horizon + 1):
        month += 1
        if month > 12:
            month = 1
            year += 1
        projected = slope * (len(revenues) - 1 + i) + intercept
        out.append((f"{year:04d}-{month:02d}", round(projected, 2)))
    return out


def sensitivity_table(base: list[tuple[str, float]]) -> list[tuple[str, float, float, float]]:
    """Return (period, low, base, high) at +/- 10% bands."""
    return [(p, round(v * 0.9, 2), v, round(v * 1.1, 2)) for p, v in base]


def write_workbook(
    actuals: list[MonthlyActual],
    forecast: list[tuple[str, float]],
    sensitivity: list[tuple[str, float, float, float]],
    out_path: Path,
) -> None:
    wb = Workbook()
    ws_actuals = wb.active
    ws_actuals.title = "Actuals"
    ws_actuals.append(["Period", "Revenue (USD)", "New Customers", "Churned Customers"])
    for a in actuals:
        ws_actuals.append([a.period, a.revenue_usd, a.new_customers, a.churned_customers])

    ws_forecast = wb.create_sheet("Forecast")
    ws_forecast.append(["Period", "Projected Revenue (USD)"])
    for p, v in forecast:
        ws_forecast.append([p, v])

    ws_sens = wb.create_sheet("Sensitivity")
    ws_sens.append(["Period", "Low (-10%)", "Base", "High (+10%)"])
    for row in sensitivity:
        ws_sens.append(list(row))

    wb.save(out_path)


def main() -> None:
    actuals = load_actuals(INPUT_CSV)
    forecast = project_forecast(actuals, FORECAST_HORIZON_MONTHS)
    sens = sensitivity_table(forecast)
    write_workbook(actuals, forecast, sens, OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX} with {len(actuals)} actuals and {len(forecast)} forecast months.")


if __name__ == "__main__":
    main()
