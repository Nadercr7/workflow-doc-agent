"""File readers: pull structured context out of .py and .xlsx without an LLM.

The LLM works much better when we hand it pre-digested structure rather
than raw bytes. These readers do the cheap, deterministic work first.
"""
from __future__ import annotations

import ast
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class PythonFileContext:
    """Pre-digested view of a Python file."""

    path: Path
    source: str
    imports: list[str]
    functions: list[str]
    classes: list[str]
    top_level_calls: list[str]

    def to_prompt_block(self) -> str:
        """Compact string suitable for embedding in an LLM prompt."""
        return (
            f"FILE: {self.path.name}\n"
            f"IMPORTS: {', '.join(self.imports) or '(none)'}\n"
            f"CLASSES: {', '.join(self.classes) or '(none)'}\n"
            f"FUNCTIONS: {', '.join(self.functions) or '(none)'}\n"
            f"TOP-LEVEL CALLS: {', '.join(self.top_level_calls) or '(none)'}\n"
            f"--- BEGIN SOURCE ---\n{self.source}\n--- END SOURCE ---\n"
        )


@dataclass(frozen=True, slots=True)
class ExcelFileContext:
    """Pre-digested view of an Excel workbook."""

    path: Path
    sheets: tuple[str, ...]
    columns_per_sheet: dict[str, list[str]]
    sample_rows: dict[str, list[list[object]]]

    def to_prompt_block(self) -> str:
        parts = [f"FILE: {self.path.name}", f"SHEETS: {', '.join(self.sheets)}"]
        for sheet in self.sheets:
            cols = self.columns_per_sheet.get(sheet, [])
            rows = self.sample_rows.get(sheet, [])
            parts.append(f"\n[Sheet: {sheet}]")
            parts.append(f"COLUMNS: {', '.join(map(str, cols)) or '(none)'}")
            if rows:
                parts.append("SAMPLE ROWS (up to 5):")
                for row in rows:
                    parts.append("  | " + " | ".join(str(c) for c in row) + " |")
        return "\n".join(parts)


def read_python_file(path: Path) -> PythonFileContext:
    """Parse a .py file with ast and return a structured context."""
    source = path.read_text(encoding="utf-8", errors="replace")
    try:
        tree = ast.parse(source, filename=str(path))
    except SyntaxError:
        # Even if we can't parse, we still hand the source to the LLM.
        return PythonFileContext(
            path=path, source=source, imports=[], functions=[], classes=[], top_level_calls=[]
        )

    imports: list[str] = []
    functions: list[str] = []
    classes: list[str] = []
    top_level_calls: list[str] = []

    for node in tree.body:
        if isinstance(node, ast.Import):
            imports.extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            for alias in node.names:
                imports.append(f"{mod}.{alias.name}" if mod else alias.name)
        elif isinstance(node, ast.FunctionDef):
            functions.append(node.name)
        elif isinstance(node, ast.AsyncFunctionDef):
            functions.append(f"async {node.name}")
        elif isinstance(node, ast.ClassDef):
            classes.append(node.name)
        elif isinstance(node, ast.Expr) and isinstance(node.value, ast.Call):
            func = node.value.func
            if isinstance(func, ast.Name):
                top_level_calls.append(func.id)
            elif isinstance(func, ast.Attribute):
                top_level_calls.append(_attr_to_str(func))

    return PythonFileContext(
        path=path,
        source=source,
        imports=imports,
        functions=functions,
        classes=classes,
        top_level_calls=top_level_calls,
    )


def _attr_to_str(node: ast.Attribute) -> str:
    """Render an ast.Attribute chain like `pd.read_csv` back to a string."""
    parts: list[str] = []
    cur: ast.expr = node
    while isinstance(cur, ast.Attribute):
        parts.append(cur.attr)
        cur = cur.value
    if isinstance(cur, ast.Name):
        parts.append(cur.id)
    return ".".join(reversed(parts))


def read_excel_file(path: Path, sample_rows: int = 5) -> ExcelFileContext:
    """Walk an .xlsx with openpyxl and return a structured context."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[str] = list(wb.sheetnames)
    columns_per_sheet: dict[str, list[str]] = {}
    sample: dict[str, list[list[object]]] = {}

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            columns_per_sheet[sheet_name] = []
            sample[sheet_name] = []
            continue

        columns_per_sheet[sheet_name] = [
            str(c) if c is not None else "" for c in header
        ]
        rows: list[list[object]] = []
        for i, row in enumerate(rows_iter):
            if i >= sample_rows:
                break
            rows.append(list(row))
        sample[sheet_name] = rows

    wb.close()
    return ExcelFileContext(
        path=path,
        sheets=tuple(sheets),
        columns_per_sheet=columns_per_sheet,
        sample_rows=sample,
    )


def discover_workflow_files(folder: Path) -> tuple[list[Path], list[Path]]:
    """Return (python_files, excel_files) inside the folder, non-recursive."""
    py = sorted(folder.glob("*.py"))
    xlsx = sorted(folder.glob("*.xlsx"))
    return py, xlsx
